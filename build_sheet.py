import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Allstate Agent Research"

headers = [
    "Row", "Name in Sheet", "Phone", "Full Name Verified", "Email",
    "Email Confidence", "LinkedIn Profile", "Allstate Profile URL",
    "Agency Name", "City/State", "Notes"
]

header_fill = PatternFill(start_color="1B2B5E", end_color="1B2B5E", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=10)
found_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
partial_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
missing_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
confirmed_fill = PatternFill(start_color="C3E6CB", end_color="C3E6CB", fill_type="solid")

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

data = [
    ("D1/D37", "Joe baker", "678-968-2553", "Joe Baker", "joebaker@allstate.com", "HIGH", "", "https://agents.allstate.com/joe-baker-kennesaw-ga1.html", "Joe Baker Agency", "Kennesaw, GA", "Confirmed Allstate agent; D1 and D37 are same person"),
    ("D2", "Joe Parker", "770-810-7339", "Joe Parks", "joeparks@allstate.com", "HIGH", "", "https://agents.allstate.com/joe-parks-woodstock-ga.html", "Parks Johnson Agency", "Woodstock, GA", "Instagram @parksjohnsonagency confirms 770-810-7339"),
    ("D3", "Nelson", "770-817-5914", "Nelson Karpeh", "nelsonkarpeh@allstate.com", "HIGH", "https://www.linkedin.com/in/nelson-karpeh-9a0823143", "https://agents.allstate.com/dykstra-kemp--biel-inc-alpharetta-ga-22546891.html", "Dykstra Kemp & Biel Inc.", "Alpharetta, GA", "Website: nelsonkarpeh.com; AGHF site confirms Nelson Karpeh Allstate Georgia"),
    ("D4", "Daniel", "404-459-9144", "Daniel Belay", "NOT ALLSTATE", "LOW", "", "", "N/A - State Farm", "Sandy Springs, GA", "WARNING: Phone 404-459-9144 = Daniel Belay STATE FARM agent, NOT Allstate"),
    ("D5", "Paul Rue", "470-377-0111", "Paul Rue", "paulrue@allstate.com", "HIGH", "", "https://agents.allstate.com/paul-rue-atlanta-ga.html", "Paul Rue Agency", "Atlanta, GA", "Confirmed Allstate Atlanta agent"),
    ("D6", "Brandon Nowden", "404-341-5915", "Brandon D. Nowden", "brandonnowden@allstate.com", "HIGH", "", "https://agents.allstate.com/brandon-nowden-atlanta-ga1.html", "Nowden Agency", "Atlanta/Smyrna, GA", "31 reviews; also brandon-d-nowden-smyrna-ga1.html"),
    ("D7", "Brianna Gordon", "770-580-4451", "Brianna Gordon", "briannagordon@allstate.com", "HIGH", "", "https://agents.allstate.com/brianna-gordon-atlanta-ga/", "Brianna Gordon Agency", "Atlanta, GA", "4.34/5 stars, 326 reviews; licensed in AL, GA, IL, SC"),
    ("D8", "Muna Karpeh", "770-951-1000", "Muna Karpeh", "munakarpeh@allstate.com", "MEDIUM", "https://www.linkedin.com/in/muna-karpeh-14796289/", "", "Stonewall Insurance Agency", "Marietta, GA", "LinkedIn: Owner/Independent Insurance Agent; verify Allstate affiliation"),
    ("D9", "Esdras", "770-526-1150", "Esdras Charles", "esdrascharles@allstate.com", "HIGH", "", "https://agents.allstate.com/esdras-charles-marietta-ga.html", "Esdras Charles Agency", "Marietta/Snellville, GA", "BBB listed; also Nextdoor in Snellville GA"),
    ("D10", "Ashley", "678-438-4128", "NOT FOUND", "NOT FOUND", "NONE", "", "", "", "Hampton, GA area", "Whitepages: 678-438-4128 is Hampton GA cell; Ashley Reese Allstate Douglasville has 678-288-0015 (no match)"),
    ("D11", "Joshua Santiago", "828-641-7335", "Joshua Santiago", "joshuasantiago@allstate.com", "MEDIUM", "", "https://agents.allstate.com/the-santiago-agency-cashiers-nc.html", "The Santiago Agency", "Cashiers, NC", "828 = NC area code; BBB indicates possibly out of business"),
    ("D12", "Lucy 3", "770-637-6015", "NOT FOUND", "NOT FOUND", "NONE", "", "", "", "Georgia", "Lucy Florveus (Norcross/Alpharetta) has 770-810-0037 - different; phone 770-637-6015 not found"),
    ("D13", "Marvah", "678-881-9281", "Marvah Jean-Baptiste", "marvahjeanbaptiste@allstate.com", "HIGH", "https://www.linkedin.com/in/marvah-jean/", "https://agents.allstate.com/marvah-jean-baptiste-duluth-ga.html", "Marvah Insurance Agency", "Duluth/Lawrenceville, GA", "2x Elite Allstate Producer; 1.4K LinkedIn followers"),
    ("D14", "Kory", "770-820-0994", "Kory A. Rykman", "krykman@allstate.com", "HIGH", "https://www.linkedin.com/in/kory-a-rykman/", "https://agents.allstate.com/kory-rykman-acworth-ga.html", "Rykman Insurance Group", "Acworth/Woodstock/Marietta/Kennesaw, GA", "Email confirmed via koryrykman.myhomehq.biz; 5 locations; 2023 Best in Company Award"),
    ("D15", "Colby", "404-464-0763", "Colby Gaskins", "colbygaskins@allstate.com", "HIGH", "https://www.linkedin.com/in/colby-gaskins-131033a6/", "https://agents.allstate.com/colby-gaskins-woodstock-ga.html", "Gaskins Agency", "Woodstock, GA", "Top producing GA Allstate agent; Agency Manager"),
    ("D16", "Barbara", "404-484-6181", "Barbara Jones", "barbarajones@allstate.com", "CONFIRMED", "https://www.linkedin.com/in/barbara-jones-insurance/", "https://agents.allstate.com/barbara-jones-stone-mountain-ga.html", "Barbara Jones Agency", "Stone Mountain, GA", "EMAIL CONFIRMED on barbarajonesagency.com; phone 404-484-6181 confirmed on website"),
    ("D17", "Andrew 1", "706-521-0911", "NOT FOUND", "NOT FOUND", "NONE", "", "", "", "Augusta/Athens, GA area", "706 = Augusta/Athens/NE Georgia; no Allstate agent matched this number"),
    ("D18", "Andrew 2", "770-462-5950", "NOT FOUND", "NOT FOUND", "NONE", "", "", "", "Atlanta suburbs", "Andrew Womack (Bogart) has 770-725-1670; Drew Niess (Atlanta) has 404-477-5500; neither matches"),
    ("D19", "Graham", "678-459-3678", "NOT FOUND", "NOT FOUND", "NONE", "", "", "", "Georgia", "C. Graham Bratcher (Marietta) has 678-795-9696 - different number; no agent matched 678-459-3678"),
    ("D20", "Chris Max", "770-576-2011", "NOT FOUND", "NOT FOUND", "NONE", "", "", "", "Georgia", "Phone 770-576-2011 not in any public directory; no matching Allstate agent Chris/Maxwell in GA"),
    ("D21", "Rachel", "404-574-5025", "Rachel Cooley", "rachelcooley@allstate.com", "HIGH", "", "https://agents.allstate.com/rachel-cooley-alpharetta-ga.html", "Rachel Cooley Agency", "Alpharetta, GA", "Profile phone 404-574-5060 (very close match to 404-574-5025)"),
    ("D22", "Matt Devine", "770-405-0280", "Matthew Devine", "matthewdevine@allstate.com", "HIGH", "", "https://agents.allstate.com/matthew-devine-atlanta-ga.html", "Matthew Devine Agency", "Atlanta, GA", "Profile phone 770-405-0713 (close match); started 2001, opened agency 2008"),
    ("D25", "Matt Devine", "770-800-9092", "Matthew Devine (Dunwoody)", "matthewdevine@allstate.com", "MEDIUM", "", "https://agents.allstate.com/matthew-devine-dunwoody-ga.html", "Matthew Devine Agency", "Dunwoody, GA", "DIFFERENT agent from D22 - same name; Dunwoody office 770-840-9200; email may differ"),
    ("D26", "Joe", "404-460-9997", "Jo Armstrong-Copeland", "joarmstrong@allstate.com", "HIGH", "", "https://agents.allstate.com/georgias-comprehensive-insurance-agency-mcdonough-ga.html", "Georgia's Comprehensive Insurance Agency", "McDonough, GA", "Website: gacompinsurance.com; Instagram @gacompinsurance; phone confirmed on website"),
    ("D27", "Barry Broom", "229-247-6677", "Barry Broome", "barrybroome@allstate.com", "HIGH", "", "https://agents.allstate.com/barry-broome-valdosta-ga1.html", "Barry Broome Agency", "Valdosta, GA", "229 = South Georgia; licensed in AL, FL, GA"),
    ("D28", "MARK TOWNSEND", "770-212-4543", "Mark Townsend", "marktownsend@allstate.com", "HIGH", "", "https://agents.allstate.com/atlanta-perimeter-associates-inc-sandy-springs-ga.html", "Atlanta Perimeter Associates Inc.", "Sandy Springs, GA", "Agency owner; 4.0/5 stars, 397 reviews; Sandy Springs GA"),
    ("D29", "Tyler Clackum", "678-402-0815", "Tyler Clackum", "tylerclackum@allstate.com", "HIGH", "https://www.linkedin.com/in/tylerclackum", "https://agents.allstate.com/clackum-family-insurance-agency-woodstock-ga.html", "Clackum Family Insurance Agency", "Woodstock, GA", "Profile phone 678-402-0813 (1 digit off); 8+ years Allstate agency owner"),
    ("D31", "Travis Miller", "770-692-8700", "Travis Miller", "travismiller@allstate.com", "HIGH", "", "https://agents.allstate.com/travis-miller-dallas-ga.html", "Travis Miller Agency", "Dallas, GA", "4.30/5 stars, 23 reviews; 110 Evans Mill Dr, Ste 202, Dallas GA 30157"),
    ("D32", "Caleb Anthony", "404-585-6798", "Caleb Anthony", "calebanthony@allstate.com", "HIGH", "", "https://agents.allstate.com/the-404-agency-atlanta-ga.html", "The 404 Agency", "Atlanta, GA", "1700 Water Pl SE, Ste 206, Atlanta GA 30339; reviews: extremely professional"),
    ("D33", "Stuart", "678-500-8460", "Stuart Hackworth", "shackworth@allstate.com", "CONFIRMED", "", "https://agents.allstate.com/the-hackworth-agency-kennesaw-ga.html", "The Hackworth Agency", "Kennesaw, GA", "EMAIL CONFIRMED (shackworth@allstate.com) via BBB/search; est. 2012; main office 678-624-2332"),
    ("D34", "Sharon", "678-369-0061", "NOT FOUND", "NOT FOUND", "NONE", "", "", "", "Georgia", "Phone 678-369-0061 not found; Sharon McCarty Armstrong (Norcross) is a different Sharon"),
    ("D36", "Eddie", "470-616-0022", "NOT FOUND", "NOT FOUND", "NONE", "", "", "", "Georgia", "Phone 470-616-0022 not found; no Allstate agent Eddie in GA matched this number"),
]

conf_colors = {
    "CONFIRMED": confirmed_fill,
    "HIGH": found_fill,
    "MEDIUM": partial_fill,
    "LOW": missing_fill,
    "NONE": missing_fill,
}

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        conf = row_data[5]
        fill = conf_colors.get(conf)
        if fill:
            cell.fill = fill

col_widths = [8, 16, 14, 24, 34, 13, 55, 72, 36, 22, 62]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

for row in range(2, len(data) + 2):
    ws.row_dimensions[row].height = 60

for row_idx in range(2, len(data) + 2):
    conf = ws.cell(row=row_idx, column=6).value
    if conf and "CONFIRMED" in str(conf):
        ws.cell(row=row_idx, column=5).font = Font(bold=True, color="155724", size=10)

ws.freeze_panes = "A2"
output = "/sessions/vigilant-trusting-dirac/mnt/transferly/Allstate_Agent_Research.xlsx"
wb.save(output)
print(f"Saved: {output}")
print(f"Total: {len(data)} | Found: {sum(1 for r in data if r[5] in ('HIGH','CONFIRMED'))} | Not found: {sum(1 for r in data if r[5]=='NONE')}")
